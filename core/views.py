from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Count, Q
from django.db import models
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import timedelta
from productos.models import Producto
from asignaciones.models import AsignacionHistorial
from core.models import Categoria, Departamento, Marca, Proveedor, EmpleadoReceptor, Ubicacion
from .error_handling import handle_errors, validate_data, BusinessLogicError, log_user_action, safe_int_conversion
from django.urls import reverse


@handle_errors(redirect_on_error='core:login')
def login_view(request):
    """Vista de login con manejo de errores"""
    try:
        if request.user.is_authenticated:
            return redirect('core:dashboard')

        if request.method == 'POST':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')

            # Validar datos de entrada
            if not username or not password:
                messages.error(request, 'Por favor ingrese usuario y contraseña')
                return render(request, 'core/login.html')

            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    log_user_action(user, "LOGIN_SUCCESS")
                    next_url = request.GET.get('next', 'core:dashboard')
                    return redirect(next_url)
                else:
                    messages.error(request, 'Su cuenta está desactivada. Contacte al administrador.')
            else:
                log_user_action(None, "LOGIN_FAILED", f"Username: {username}")
                messages.error(request, 'Usuario o contraseña incorrectos')

        return render(request, 'core/login.html')

    except Exception as e:
        # El decorador manejará la excepción
        raise


@handle_errors(redirect_on_error='core:login')
def logout_view(request):
    """Vista de logout con manejo de errores"""
    try:
        if request.user.is_authenticated:
            log_user_action(request.user, "LOGOUT")
        logout(request)
        messages.success(request, 'Has cerrado sesión correctamente')
        return redirect('core:login')
    except Exception as e:
        raise


@login_required
@handle_errors(redirect_on_error='core:login')
def dashboard(request):
    """Dashboard principal con resumen del inventario - Con manejo de errores robusto"""
    try:
        # Intentar obtener estadísticas con valores por defecto en caso de error
        context = {
            'total_productos': 0,
            'productos_disponibles': 0,
            'productos_entregados': 0,
            'productos_averiados': 0,
            'total_departamentos': 0,
            'total_empleados': 0,
            'productos_por_departamento': [],
            'empleados_con_productos': [],
            'productos_por_categoria': [],
            'asignaciones_recientes': 0,
            'productos_garantia_venciendo': [],
            'asignaciones_pendientes': []
        }

        # Estadísticas generales de productos con manejo de errores
        try:
            context['total_productos'] = Producto.objects.filter(activo=True).count()
            context['productos_disponibles'] = Producto.objects.filter(estado='DISPONIBLE', activo=True).count()
            context['productos_entregados'] = Producto.objects.filter(estado='ENTREGADO', activo=True).count()
            context['productos_averiados'] = Producto.objects.filter(estado='AVERIADO', activo=True).count()
        except Exception as e:
            messages.warning(request, 'Error al cargar estadísticas de productos')
            log_user_action(request.user, "DASHBOARD_ERROR_PRODUCTOS", str(e))

        # Estadísticas de departamentos y empleados
        try:
            context['total_departamentos'] = Departamento.objects.filter(activo=True).count()
            context['total_empleados'] = EmpleadoReceptor.objects.filter(activo=True).count()
        except Exception as e:
            messages.warning(request, 'Error al cargar estadísticas de departamentos/empleados')
            log_user_action(request.user, "DASHBOARD_ERROR_DEPT_EMP", str(e))

        # Departamentos con más productos asignados
        try:
            context['productos_por_departamento'] = Departamento.objects.annotate(
                total_entregados=Count('asignacionhistorial__producto',
                                     filter=Q(asignacionhistorial__fecha_devolucion__isnull=True,
                                             asignacionhistorial__tipo_asignacion__in=['ENTREGA', 'PRESTAMO']))
            ).filter(activo=True).order_by('-total_entregados')[:10]
        except Exception as e:
            messages.warning(request, 'Error al cargar productos por departamento')

        # Empleados con más productos asignados
        try:
            context['empleados_con_productos'] = EmpleadoReceptor.objects.annotate(
                total_productos=Count('asignacionhistorial__producto',
                                    filter=Q(asignacionhistorial__fecha_devolucion__isnull=True,
                                            asignacionhistorial__tipo_asignacion__in=['ENTREGA', 'PRESTAMO']))
            ).filter(activo=True, total_productos__gt=0).order_by('-total_productos')[:10]
        except Exception as e:
            messages.warning(request, 'Error al cargar empleados con productos')

        # Productos por categoría
        try:
            context['productos_por_categoria'] = Categoria.objects.annotate(
                total=Count('producto', filter=Q(producto__activo=True)),
                disponibles=Count('producto', filter=Q(producto__estado='DISPONIBLE', producto__activo=True)),
                entregados=Count('producto', filter=Q(producto__estado='ENTREGADO', producto__activo=True))
            ).filter(activo=True)
        except Exception as e:
            messages.warning(request, 'Error al cargar productos por categoría')

        # Asignaciones recientes (últimos 30 días)
        try:
            fecha_limite = timezone.now().date() - timedelta(days=30)
            context['asignaciones_recientes'] = AsignacionHistorial.objects.filter(
                fecha_entrega__gte=fecha_limite,
                activo=True
            ).count()
        except Exception as e:
            messages.warning(request, 'Error al cargar asignaciones recientes')

        # Productos con garantía próxima a vencer (30 días)
        try:
            fecha_limite_garantia = timezone.now().date() + timedelta(days=30)
            context['productos_garantia_venciendo'] = Producto.objects.filter(
                fecha_fin_garantia__lte=fecha_limite_garantia,
                fecha_fin_garantia__gte=timezone.now().date(),
                activo=True
            ).order_by('fecha_fin_garantia')[:10]
        except Exception as e:
            messages.warning(request, 'Error al cargar productos con garantía próxima a vencer')

        # Asignaciones pendientes de confirmación
        try:
            context['asignaciones_pendientes'] = AsignacionHistorial.objects.filter(
                confirmado_empleado=False,
                fecha_devolucion__isnull=True,
                activo=True
            ).order_by('-fecha_entrega')[:10]
        except Exception as e:
            messages.warning(request, 'Error al cargar asignaciones pendientes')

        # Completar el contexto original
        context.update({
            'productos_por_departamento': context['productos_por_departamento'],
            'empleados_con_productos': context['empleados_con_productos'],
            'productos_por_categoria': context['productos_por_categoria'],
            'productos_garantia_venciendo': context['productos_garantia_venciendo'],
            'asignaciones_pendientes': context['asignaciones_pendientes'],
        })

        log_user_action(request.user, "DASHBOARD_ACCESSED")
        return render(request, 'core/dashboard.html', context)

    except Exception as e:
        # El decorador manejará la excepción
        raise


@login_required
def inventario_por_categoria(request, categoria_id=None):
    """Vista para mostrar productos por categoría"""
    categorias = Categoria.objects.filter(activo=True)

    if categoria_id:
        categoria_seleccionada = Categoria.objects.get(id=categoria_id)
        productos = Producto.objects.filter(
            categoria=categoria_seleccionada,
            activo=True
        ).order_by('-created_at')
    else:
        categoria_seleccionada = None
        productos = Producto.objects.filter(activo=True).order_by('-created_at')

    # Filtros adicionales
    estado_filtro = request.GET.get('estado')
    if estado_filtro:
        productos = productos.filter(estado=estado_filtro)

    context = {
        'categorias': categorias,
        'categoria_seleccionada': categoria_seleccionada,
        'productos': productos,
        'estados': Producto.ESTADOS,
        'estado_filtro': estado_filtro,
    }

    return render(request, 'productos/inventario_categoria.html', context)


@login_required
def inventario_por_departamento(request, departamento_id=None):
    """Vista para mostrar productos asignados por departamento"""
    departamentos = Departamento.objects.filter(activo=True)

    if departamento_id:
        departamento_seleccionado = Departamento.objects.get(id=departamento_id)
        asignaciones = AsignacionHistorial.objects.filter(
            departamento=departamento_seleccionado,
            fecha_devolucion__isnull=True,
            activo=True
        ).order_by('-fecha_entrega')
    else:
        departamento_seleccionado = None
        asignaciones = AsignacionHistorial.objects.filter(
            fecha_devolucion__isnull=True,
            activo=True
        ).order_by('-fecha_entrega')

    context = {
        'departamentos': departamentos,
        'departamento_seleccionado': departamento_seleccionado,
        'asignaciones': asignaciones,
    }

    return render(request, 'asignaciones/inventario_departamento.html', context)


@login_required
def admin_panel(request):
    """Panel de administración personalizado con estadísticas precisas"""
    from django.contrib.auth.models import User
    from productos.models import Producto
    from asignaciones.models import AsignacionHistorial
    from core.models import Categoria, Departamento, Marca, Proveedor, EmpleadoReceptor, Ubicacion

    # Estadísticas del sistema - datos precisos
    total_usuarios = User.objects.count()
    total_productos = Producto.objects.count()  # Todos los productos
    productos_disponibles = Producto.objects.filter(estado='DISPONIBLE').count()
    productos_entregados = Producto.objects.filter(estado='ENTREGADO').count()
    productos_averiados = Producto.objects.filter(estado='AVERIADO').count()
    productos_rotos = Producto.objects.filter(estado='ROTO').count()
    productos_recogidos = Producto.objects.filter(estado='RECOGIDO').count()

    # Asignaciones - solo las activas (sin devolución)
    total_asignaciones = AsignacionHistorial.objects.filter(
        fecha_devolucion__isnull=True,
        tipo_asignacion='ENTREGA'
    ).count()

    # Datos maestros
    categorias = Categoria.objects.all()
    departamentos = Departamento.objects.all()
    total_marcas = Marca.objects.count()
    total_empleados = EmpleadoReceptor.objects.count()
    total_proveedores = Proveedor.objects.count()
    total_ubicaciones = Ubicacion.objects.count()

    # Actividad reciente (últimos productos agregados)
    productos_recientes = Producto.objects.select_related(
        'categoria', 'marca', 'proveedor'
    ).order_by('-created_at')[:10]

    # Asignaciones recientes (últimas 10)
    asignaciones_recientes = AsignacionHistorial.objects.select_related(
        'producto', 'empleado_receptor', 'departamento'
    ).filter(tipo_asignacion='ENTREGA').order_by('-fecha_entrega')[:10]

    # Empleados recientes (últimos registrados)
    empleados_recientes = EmpleadoReceptor.objects.select_related(
        'departamento'
    ).order_by('-created_at')[:10]

    # Marcas recientes (últimas agregadas)
    marcas_recientes = Marca.objects.order_by('-created_at')[:10]

    # Estadísticas por estado para verificar que suman correctamente
    estadisticas_estados = {
        'DISPONIBLE': productos_disponibles,
        'ENTREGADO': productos_entregados,
        'AVERIADO': productos_averiados,
        'ROTO': productos_rotos,
        'RECOGIDO': productos_recogidos,
    }

    # Verificación de integridad de datos
    suma_estados = sum(estadisticas_estados.values())
    productos_sin_estado = total_productos - suma_estados

    context = {
        # Estadísticas principales
        'total_usuarios': total_usuarios,
        'total_productos': total_productos,
        'productos_disponibles': productos_disponibles,
        'productos_entregados': productos_entregados,
        'productos_averiados': productos_averiados,
        'productos_rotos': productos_rotos,
        'productos_recogidos': productos_recogidos,
        'total_asignaciones': total_asignaciones,

        # Datos maestros
        'categorias': categorias,
        'departamentos': departamentos,
        'total_marcas': total_marcas,
        'total_empleados': total_empleados,
        'total_proveedores': total_proveedores,
        'total_ubicaciones': total_ubicaciones,

        # Actividad reciente
        'productos_recientes': productos_recientes,
        'asignaciones_recientes': asignaciones_recientes,
        'empleados_recientes': empleados_recientes,
        'marcas_recientes': marcas_recientes,

        # Estadísticas detalladas
        'estadisticas_estados': estadisticas_estados,
        'suma_estados': suma_estados,
        'productos_sin_estado': productos_sin_estado,

        # Información del sistema
        'fecha_actualizacion': timezone.now(),
    }

    return render(request, 'core/admin_panel.html', context)


@login_required
@handle_errors()
def gestionar_categorias(request):
    """Vista para gestionar categorías"""
    from django.db.models import Count
    import json

    # Obtener categorías con el conteo de productos
    categorias = Categoria.objects.annotate(
        productos_count=Count('producto', filter=models.Q(producto__activo=True))
    ).order_by('nombre')

    # Calcular estadísticas
    total_categorias = categorias.count()
    categorias_activas = categorias.filter(activo=True).count()
    categorias_inactivas = categorias.filter(activo=False).count()
    categorias_con_productos = categorias.filter(productos_count__gt=0).count()

    if request.method == 'POST':
        action = request.POST.get('action')
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        codigo = request.POST.get('codigo', '').strip()

        if action == 'crear':
            # Procesar campos específicos
            campos_especificos = {}
            campo_nombres = request.POST.getlist('campo_nombre[]')
            campo_tipos = request.POST.getlist('campo_tipo[]')
            campo_obligatorios = request.POST.getlist('campo_obligatorio[]')
            campo_opciones = request.POST.getlist('campo_opciones[]')

            for i, campo_nombre in enumerate(campo_nombres):
                if campo_nombre.strip():
                    campo_config = {
                        'tipo': campo_tipos[i] if i < len(campo_tipos) else 'text',
                        'obligatorio': str(i) in campo_obligatorios,
                        'label': campo_nombre.strip()
                    }

                    # Si es select o radio, agregar opciones
                    if campo_config['tipo'] in ['select', 'radio'] and i < len(campo_opciones):
                        opciones = [opt.strip() for opt in campo_opciones[i].split(',') if opt.strip()]
                        if opciones:
                            campo_config['opciones'] = opciones

                    campos_especificos[f'campo_{i+1}'] = campo_config

            # Generar código único si no se proporciona
            if not codigo:
                base_codigo = ''.join([c for c in nombre if c.isalnum()])[:6].upper()
                codigo = base_codigo
                contador = 1
                while Categoria.objects.filter(codigo=codigo).exists():
                    codigo = f"{base_codigo}{contador}"
                    contador += 1

            categoria = Categoria(
                nombre=nombre,
                descripcion=descripcion,
                codigo=codigo,
                usuario_creacion=request.user
            )
            categoria.campos_especificos = campos_especificos
            categoria.save()
            messages.success(request, f'Categoría "{nombre}" creada correctamente con {len(campos_especificos)} características personalizadas.')
            return redirect('core:gestionar_categorias')

        elif action == 'editar':
            categoria_id = request.POST.get('categoria_id')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')

            # Procesar campos específicos para edición
            campos_especificos = {}
            campo_nombres = request.POST.getlist('campo_nombre[]')
            campo_tipos = request.POST.getlist('campo_tipo[]')
            campo_obligatorios = request.POST.getlist('campo_obligatorio[]')
            campo_opciones = request.POST.getlist('campo_opciones[]')

            for i, campo_nombre in enumerate(campo_nombres):
                if campo_nombre.strip():
                    campo_config = {
                        'tipo': campo_tipos[i] if i < len(campo_tipos) else 'text',
                        'obligatorio': str(i) in campo_obligatorios,
                        'label': campo_nombre.strip()
                    }

                    if campo_config['tipo'] in ['select', 'radio'] and i < len(campo_opciones):
                        opciones = [opt.strip() for opt in campo_opciones[i].split(',') if opt.strip()]
                        if opciones:
                            campo_config['opciones'] = opciones

                    campos_especificos[f'campo_{i+1}'] = campo_config

            try:
                categoria = Categoria.objects.get(id=categoria_id)
                categoria.nombre = nombre
                categoria.descripcion = descripcion
                categoria.campos_especificos = campos_especificos
                categoria.save()
                messages.success(request, f'Categoría "{nombre}" actualizada exitosamente.')
            except Categoria.DoesNotExist:
                messages.error(request, 'Categoría no encontrada.')

        elif action == 'toggle_activo':
            categoria_id = request.POST.get('categoria_id')
            try:
                categoria = Categoria.objects.get(id=categoria_id)
                categoria.activo = not categoria.activo
                categoria.save()
                estado = 'activada' if categoria.activo else 'desactivada'
                messages.success(request, f'Categoría "{categoria.nombre}" {estado}.')
            except Categoria.DoesNotExist:
                messages.error(request, 'Categoría no encontrada.')

        return redirect('core:gestionar_categorias')

    context = {
        'categorias': categorias,
        'total_categorias': total_categorias,
        'categorias_activas': categorias_activas,
        'categorias_inactivas': categorias_inactivas,
        'categorias_con_productos': categorias_con_productos,
    }
    return render(request, 'core/gestionar_categorias.html', context)


@login_required
def gestionar_departamentos(request):
    """Vista para gestionar departamentos"""
    from django.db.models import Count

    # Obtener departamentos con el conteo de asignaciones y empleados
    departamentos = Departamento.objects.annotate(
        asignaciones_count=Count('asignacionhistorial', filter=Q(
            asignacionhistorial__fecha_devolucion__isnull=True,
            asignacionhistorial__activo=True
        )),
        empleados_count=Count('empleadoreceptor', filter=Q(empleadoreceptor__activo=True))
    ).order_by('nombre')

    # Calcular estadísticas
    total_departamentos = departamentos.count()
    departamentos_activos = departamentos.filter(activo=True).count()
    departamentos_inactivos = departamentos.filter(activo=False).count()
    departamentos_con_asignaciones = departamentos.filter(asignaciones_count__gt=0).count()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear':
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')
            codigo = request.POST.get('codigo', '').strip()
            # Generar código único si no se proporciona
            if not codigo:
                base_codigo = ''.join([c for c in nombre if c.isalnum()])[:6].upper()
                codigo = base_codigo
                contador = 1
                while Departamento.objects.filter(codigo=codigo).exists():
                    codigo = f"{base_codigo}{contador}"
                    contador += 1
            if nombre:
                Departamento.objects.create(
                    nombre=nombre,
                    descripcion=descripcion,
                    codigo=codigo,
                    usuario_creacion=request.user
                )
                messages.success(request, f'Departamento "{nombre}" creado exitosamente.')
            else:
                messages.error(request, 'El nombre del departamento es requerido.')

        elif action == 'editar':
            departamento_id = request.POST.get('departamento_id')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')

            try:
                departamento = Departamento.objects.get(id=departamento_id)
                departamento.nombre = nombre
                departamento.descripcion = descripcion
                departamento.save()
                messages.success(request, f'Departamento "{nombre}" actualizado exitosamente.')
            except Departamento.DoesNotExist:
                messages.error(request, 'Departamento no encontrado.')

        elif action == 'toggle_activo':
            departamento_id = request.POST.get('departamento_id')
            try:
                departamento = Departamento.objects.get(id=departamento_id)
                departamento.activo = not departamento.activo
                departamento.save()
                estado = 'activado' if departamento.activo else 'desactivado'
                messages.success(request, f'Departamento "{departamento.nombre}" {estado}.')
            except Departamento.DoesNotExist:
                messages.error(request, 'Departamento no encontrado.')

        elif action == 'eliminar':
            departamento_id = request.POST.get('departamento_id')
            try:
                departamento = Departamento.objects.get(id=departamento_id)
                nombre = departamento.nombre
                departamento.delete()
                messages.success(request, f'Departamento "{nombre}" eliminado correctamente.')
            except Departamento.DoesNotExist:
                messages.error(request, 'Departamento no encontrado.')

        return redirect('core:gestionar_departamentos')

    context = {
        'departamentos': departamentos,
        'total_departamentos': total_departamentos,
        'departamentos_activos': departamentos_activos,
        'departamentos_inactivos': departamentos_inactivos,
        'departamentos_con_asignaciones': departamentos_con_asignaciones,
    }
    return render(request, 'core/gestionar_departamentos.html', context)


@login_required
def gestionar_marcas(request):
    """Vista para gestionar marcas"""
    marcas = Marca.objects.all().order_by('nombre')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear':
            nombre = request.POST.get('nombre')
            if nombre:
                Marca.objects.create(nombre=nombre)
                messages.success(request, f'Marca "{nombre}" creada exitosamente.')
            else:
                messages.error(request, 'El nombre de la marca es requerido.')

        elif action == 'editar':
            marca_id = request.POST.get('marca_id')
            nombre = request.POST.get('nombre')
            try:
                marca = Marca.objects.get(id=marca_id)
                marca.nombre = nombre
                marca.save()
                messages.success(request, f'Marca "{nombre}" actualizada exitosamente.')
            except Marca.DoesNotExist:
                messages.error(request, 'Marca no encontrada.')

        elif action == 'toggle_activo':
            marca_id = request.POST.get('marca_id')
            try:
                marca = Marca.objects.get(id=marca_id)
                marca.activo = not marca.activo
                marca.save()
                estado = 'activada' if marca.activo else 'desactivada'
                messages.success(request, f'Marca "{marca.nombre}" {estado}.')
            except Marca.DoesNotExist:
                messages.error(request, 'Marca no encontrada.')

        elif action == 'eliminar':
            marca_id = request.POST.get('marca_id')
            try:
                marca = Marca.objects.get(id=marca_id)
                marca.delete()
                messages.success(request, f'Marca "{marca.nombre}" eliminada exitosamente.')
            except Marca.DoesNotExist:
                messages.error(request, 'Marca no encontrada.')

        return redirect('core:gestionar_marcas')

    context = {
        'marcas': marcas,
    }
    return render(request, 'core/gestionar_marcas.html', context)


@login_required
def gestionar_empleados(request):
    """Vista para gestionar empleados receptores"""
    from django.db.models import Count

    # Obtener empleados con el conteo de productos asignados
    empleados = EmpleadoReceptor.objects.annotate(
        productos_asignados_count=Count('asignacionhistorial', filter=Q(
            asignacionhistorial__fecha_devolucion__isnull=True,
            asignacionhistorial__activo=True
        ))
    ).select_related('departamento').order_by('nombre')

    departamentos = Departamento.objects.filter(activo=True)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear':
            nombre = request.POST.get('nombre')
            dni = request.POST.get('dni')
            email = request.POST.get('email')
            telefono = request.POST.get('telefono')
            departamento_id = request.POST.get('departamento')
            puesto = request.POST.get('puesto')
            fecha_alta = request.POST.get('fecha_alta')

            if nombre and dni and email and departamento_id and fecha_alta:
                try:
                    departamento = Departamento.objects.get(id=departamento_id)
                    EmpleadoReceptor.objects.create(
                        nombre=nombre,
                        dni=dni,
                        email=email,
                        telefono=telefono,
                        departamento=departamento,
                        puesto=puesto,
                        fecha_alta=fecha_alta
                    )
                    messages.success(request, f'Empleado "{nombre}" creado exitosamente.')
                except Exception as e:
                    messages.error(request, f'Error al crear empleado: {str(e)}')
            else:
                messages.error(request, 'Todos los campos obligatorios son requeridos.')

        elif action == 'toggle_activo':
            empleado_id = request.POST.get('empleado_id')
            try:
                empleado = EmpleadoReceptor.objects.get(id=empleado_id)
                empleado.activo = not empleado.activo
                empleado.save()
                estado = 'activado' if empleado.activo else 'desactivado'
                messages.success(request, f'Empleado "{empleado.nombre}" {estado}.')
            except EmpleadoReceptor.DoesNotExist:
                messages.error(request, 'Empleado no encontrado.')

        elif action == 'eliminar':
            empleado_id = request.POST.get('empleado_id')
            try:
                empleado = EmpleadoReceptor.objects.get(id=empleado_id)
                empleado.delete()
                messages.success(request, f'Empleado "{empleado.nombre}" eliminado exitosamente.')
            except EmpleadoReceptor.DoesNotExist:
                messages.error(request, 'Empleado no encontrado.')

        elif action == 'editar_guardar':
            empleado_id = request.POST.get('empleado_id')
            nombre = request.POST.get('nombre')
            dni = request.POST.get('dni')
            email = request.POST.get('email')
            telefono = request.POST.get('telefono')
            departamento_id = request.POST.get('departamento')
            puesto = request.POST.get('puesto')
            fecha_alta = request.POST.get('fecha_alta')
            try:
                empleado = EmpleadoReceptor.objects.get(id=empleado_id)
                empleado.nombre = nombre
                empleado.dni = dni
                empleado.email = email
                empleado.telefono = telefono
                empleado.departamento = Departamento.objects.get(id=departamento_id)
                empleado.puesto = puesto
                empleado.fecha_alta = fecha_alta
                empleado.save()
                messages.success(request, f'Datos de "{empleado.nombre}" actualizados correctamente.')
            except EmpleadoReceptor.DoesNotExist:
                messages.error(request, 'Empleado no encontrado.')
            except Exception as e:
                messages.error(request, f'Error al actualizar empleado: {str(e)}')

        return redirect('core:gestionar_empleados')

    context = {
        'empleados': empleados,
        'departamentos': departamentos,
    }
    return render(request, 'core/gestionar_empleados.html', context)


@login_required
def gestionar_proveedores(request):
    """Vista para gestionar proveedores"""
    proveedores = Proveedor.objects.all().order_by('nombre')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear':
            nombre = request.POST.get('nombre')
            nif_cif = request.POST.get('nif_cif')
            telefono = request.POST.get('telefono')
            email = request.POST.get('email')
            direccion = request.POST.get('direccion')

            if nombre and nif_cif:
                try:
                    Proveedor.objects.create(
                        nombre=nombre,
                        nif_cif=nif_cif,
                        telefono=telefono,
                        email=email,
                        direccion=direccion
                    )
                    messages.success(request, f'Proveedor "{nombre}" creado exitosamente.')
                except Exception as e:
                    messages.error(request, f'Error al crear proveedor: {str(e)}')
            else:
                messages.error(request, 'El nombre y NIF/CIF son requeridos.')

        elif action == 'toggle_activo':
            proveedor_id = request.POST.get('proveedor_id')
            try:
                proveedor = Proveedor.objects.get(id=proveedor_id)
                proveedor.activo = not proveedor.activo
                proveedor.save()
                estado = 'activado' if proveedor.activo else 'desactivado'
                messages.success(request, f'Proveedor "{proveedor.nombre}" {estado}.')
            except Proveedor.DoesNotExist:
                messages.error(request, 'Proveedor no encontrado.')

        elif action == 'eliminar':
            proveedor_id = request.POST.get('proveedor_id')
            try:
                proveedor = Proveedor.objects.get(id=proveedor_id)
                proveedor.delete()
                messages.success(request, f'Proveedor "{proveedor.nombre}" eliminado exitosamente.')
            except Proveedor.DoesNotExist:
                messages.error(request, 'Proveedor no encontrado.')

        return redirect('core:gestionar_proveedores')

    context = {
        'proveedores': proveedores,
    }
    return render(request, 'core/gestionar_proveedores.html', context)


@login_required
def gestionar_ubicaciones(request):
    """Vista para gestionar ubicaciones"""
    ubicaciones = Ubicacion.objects.all().order_by('edificio', 'planta', 'sala')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear':
            nombre = request.POST.get('nombre')
            edificio = request.POST.get('edificio')
            planta = request.POST.get('planta')
            sala = request.POST.get('sala')
            descripcion = request.POST.get('descripcion')

            if nombre:
                Ubicacion.objects.create(
                    nombre=nombre,
                    edificio=edificio,
                    planta=planta,
                    sala=sala,
                    descripcion=descripcion
                )
                messages.success(request, f'Ubicación "{nombre}" creada exitosamente.')
            else:
                messages.error(request, 'El nombre de la ubicación es requerido.')

        elif action == 'toggle_activo':
            ubicacion_id = request.POST.get('ubicacion_id')
            try:
                ubicacion = Ubicacion.objects.get(id=ubicacion_id)
                ubicacion.activo = not ubicacion.activo
                ubicacion.save()
                estado = 'activada' if ubicacion.activo else 'desactivada'
                messages.success(request, f'Ubicación "{ubicacion.nombre}" {estado}.')
            except Ubicacion.DoesNotExist:
                messages.error(request, 'Ubicación no encontrada.')

        return redirect('core:gestionar_ubicaciones')

    context = {
        'ubicaciones': ubicaciones,
    }
    return render(request, 'core/gestionar_ubicaciones.html', context)


# ==========================================
# HERRAMIENTAS DEL SISTEMA
# ==========================================

@login_required
def generar_backup(request):
    """Genera un backup completo del sistema"""
    import json
    import os
    from django.core import serializers
    from django.http import JsonResponse, HttpResponse
    from django.core.management import call_command
    from io import StringIO
    from datetime import datetime
    import zipfile
    from django.conf import settings
    import tempfile

    try:
        # Crear nombre del archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_sistema_{timestamp}"

        if request.method == 'POST':
            # Generar backup de la base de datos
            backup_data = {}

            # Serializar todos los modelos
            from productos.models import Producto
            from asignaciones.models import AsignacionHistorial
            from productos.models_pegatinas import TipoProducto, DatosPersonalizados

            models_to_backup = [
                ('categorias', Categoria.objects.all()),
                ('departamentos', Departamento.objects.all()),
                ('marcas', Marca.objects.all()),
                ('proveedores', Proveedor.objects.all()),
                ('empleados', EmpleadoReceptor.objects.all()),
                ('ubicaciones', Ubicacion.objects.all()),
                ('productos', Producto.objects.all()),
                ('asignaciones', AsignacionHistorial.objects.all()),
                ('tipos_productos', TipoProducto.objects.all()),
                ('datos_personalizados', DatosPersonalizados.objects.all()),
            ]

            for model_name, queryset in models_to_backup:
                backup_data[model_name] = json.loads(serializers.serialize('json', queryset))

            # Crear archivo ZIP con el backup
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{backup_filename}.zip"'

            # Crear archivo temporal
            with tempfile.NamedTemporaryFile() as tmp_file:
                with zipfile.ZipFile(tmp_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Agregar datos JSON
                    json_data = json.dumps(backup_data, indent=2, ensure_ascii=False, default=str)
                    zipf.writestr(f"{backup_filename}.json", json_data)

                    # Agregar metadatos del backup
                    metadata = {
                        'fecha_backup': datetime.now().isoformat(),
                        'total_productos': Producto.objects.count(),
                        'total_asignaciones': AsignacionHistorial.objects.count(),
                        'version_sistema': '1.0',
                        'usuario_backup': request.user.username
                    }
                    zipf.writestr("metadata.json", json.dumps(metadata, indent=2))

                # Leer el contenido del archivo temporal
                tmp_file.seek(0)
                response.write(tmp_file.read())

            messages.success(request, f'Backup generado exitosamente: {backup_filename}.zip')
            return response

        # GET request - mostrar información
        stats = {
            'total_productos': Producto.objects.count(),
            'total_asignaciones': AsignacionHistorial.objects.count(),
            'total_categorias': Categoria.objects.count(),
            'total_empleados': EmpleadoReceptor.objects.count(),
            'backup_filename': backup_filename
        }

        return JsonResponse({'success': True, 'stats': stats})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def exportar_datos(request):
    """Exporta datos del sistema en diferentes formatos"""
    import csv
    import json
    from django.http import HttpResponse
    from datetime import datetime
    from productos.models import Producto
    from asignaciones.models import AsignacionHistorial

    formato = request.GET.get('formato', 'excel')
    tipo_datos = request.GET.get('tipo', 'productos')

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    try:
        if tipo_datos == 'productos':
            productos = Producto.objects.select_related('categoria', 'marca', 'proveedor', 'ubicacion_actual').all()

            if formato == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="productos_{timestamp}.csv"'

                writer = csv.writer(response)
                writer.writerow([
                    'Número de Serie', 'Categoría', 'Marca', 'Modelo', 'Estado', 'Condición',
                    'Precio Compra', 'Fecha Compra', 'Proveedor', 'Ubicación', 'Garantía Hasta'
                ])

                for producto in productos:
                    writer.writerow([
                        producto.numero_serie,
                        producto.categoria.nombre,
                        producto.marca.nombre,
                        producto.modelo,
                        producto.get_estado_display(),
                        producto.get_condicion_display(),
                        producto.precio_compra,
                        producto.fecha_compra,
                        producto.proveedor.nombre if producto.proveedor else '',
                        producto.ubicacion_actual.nombre if producto.ubicacion_actual else '',
                        producto.fecha_fin_garantia or ''
                    ])

                return response

            elif formato == 'json':
                response = HttpResponse(content_type='application/json; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="productos_{timestamp}.json"'

                productos_data = []
                for producto in productos:
                    productos_data.append({
                        'numero_serie': producto.numero_serie,
                        'categoria': producto.categoria.nombre,
                        'marca': producto.marca.nombre,
                        'modelo': producto.modelo,
                        'estado': producto.estado,
                        'condicion': producto.condicion,
                        'precio_compra': float(producto.precio_compra),
                        'fecha_compra': producto.fecha_compra.isoformat(),
                        'proveedor': producto.proveedor.nombre if producto.proveedor else None,
                        'ubicacion': producto.ubicacion_actual.nombre if producto.ubicacion_actual else None,
                        'garantia_hasta': producto.fecha_fin_garantia.isoformat() if producto.fecha_fin_garantia else None
                    })

                response.write(json.dumps(productos_data, indent=2, ensure_ascii=False))
                return response

            else:  # Excel por defecto
                try:
                    import openpyxl
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Font, PatternFill

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Productos"

                    # Encabezados
                    headers = [
                        'Número de Serie', 'Categoría', 'Marca', 'Modelo', 'Estado', 'Condición',
                        'Precio Compra', 'Fecha Compra', 'Proveedor', 'Ubicación', 'Garantía Hasta'
                    ]

                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                    # Datos
                    for row, producto in enumerate(productos, 2):
                        ws.cell(row=row, column=1, value=producto.numero_serie)
                        ws.cell(row=row, column=2, value=producto.categoria.nombre)
                        ws.cell(row=row, column=3, value=producto.marca.nombre)
                        ws.cell(row=row, column=4, value=producto.modelo)
                        ws.cell(row=row, column=5, value=producto.get_estado_display())
                        ws.cell(row=row, column=6, value=producto.get_condicion_display())
                        ws.cell(row=row, column=7, value=float(producto.precio_compra))
                        ws.cell(row=row, column=8, value=producto.fecha_compra)
                        ws.cell(row=row, column=9, value=producto.proveedor.nombre if producto.proveedor else '')
                        ws.cell(row=row, column=10, value=producto.ubicacion_actual.nombre if producto.ubicacion_actual else '')
                        ws.cell(row=row, column=11, value=producto.fecha_fin_garantia)

                    # Ajustar ancho de columnas
                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col)].width = 15

                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename="productos_{timestamp}.xlsx"'
                    wb.save(response)

                    return response

                except ImportError:
                    messages.error(request, 'openpyxl no está instalado. Usando formato CSV como alternativa.')
                    return exportar_datos(request.GET.copy().update({'formato': 'csv'}))

        elif tipo_datos == 'asignaciones':
            asignaciones = AsignacionHistorial.objects.select_related(
                'producto', 'empleado_receptor', 'departamento'
            ).all()

            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="asignaciones_{timestamp}.csv"'

            writer = csv.writer(response)
            writer.writerow([
                'Producto', 'Número Serie', 'Empleado', 'Departamento', 'Tipo Asignación',
                'Fecha Entrega', 'Fecha Devolución', 'Estado', 'Observaciones'
            ])

            for asignacion in asignaciones:
                writer.writerow([
                    f"{asignacion.producto.marca.nombre} {asignacion.producto.modelo}",
                    asignacion.producto.numero_serie,
                    asignacion.empleado_receptor.nombre if asignacion.empleado_receptor else '',
                    asignacion.departamento.nombre,
                    asignacion.get_tipo_asignacion_display(),
                    asignacion.fecha_entrega,
                    asignacion.fecha_devolucion or '',
                    'Activa' if not asignacion.fecha_devolucion else 'Devuelta',
                    asignacion.observaciones or ''
                ])

            return response

    except Exception as e:
        messages.error(request, f'Error al exportar datos: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def limpiar_cache(request):
    """Limpia la caché del sistema"""
    from django.core.cache import cache
    from django.http import JsonResponse
    import os
    import shutil

    try:
        if request.method == 'POST':
            # Limpiar caché de Django
            cache.clear()

            # Limpiar archivos temporales
            temp_dirs = [
                '/tmp/django_cache',
                os.path.join(os.getcwd(), 'tmp'),
                os.path.join(os.getcwd(), 'cache'),
            ]

            archivos_eliminados = 0
            for temp_dir in temp_dirs:
                if os.path.exists(temp_dir):
                    try:
                        shutil.rmtree(temp_dir)
                        archivos_eliminados += 1
                    except:
                        pass

            # Limpiar sesiones expiradas
            from django.core.management import call_command
            call_command('clearsessions')

            messages.success(request, f'Caché limpiado exitosamente. Se eliminaron {archivos_eliminados} directorios temporales.')

            return JsonResponse({
                'success': True,
                'message': 'Caché limpiado exitosamente',
                'archivos_eliminados': archivos_eliminados
            })

        return JsonResponse({'success': False, 'error': 'Método no permitido'})

    except Exception as e:
        messages.error(request, f'Error al limpiar caché: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def generar_reporte_inventario(request):
    """Genera un reporte completo del inventario"""
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    from datetime import datetime
    from django.db.models import Count, Sum, Avg, Q
    from productos.models import Producto
    from asignaciones.models import AsignacionHistorial

    try:
        # Recopilar estadísticas
        total_productos = Producto.objects.count()
        productos_por_estado = Producto.objects.values('estado').annotate(count=Count('id')).order_by('-count')

        productos_por_categoria = Categoria.objects.annotate(
            total=Count('producto'),
            disponibles=Count('producto', filter=Q(producto__estado='DISPONIBLE')),
            entregados=Count('producto', filter=Q(producto__estado='ENTREGADO'))
        ).filter(total__gt=0).order_by('-total')

        productos_por_departamento = Departamento.objects.annotate(
            total_asignados=Count('asignacionhistorial__producto',
                                filter=Q(asignacionhistorial__fecha_devolucion__isnull=True,
                                        asignacionhistorial__tipo_asignacion='ENTREGA'))
        ).filter(total_asignados__gt=0).order_by('-total_asignados')

        valor_total_inventario = Producto.objects.aggregate(
            total=Sum('precio_compra')
        )['total'] or 0

        precio_promedio = Producto.objects.aggregate(
            promedio=Avg('precio_compra')
        )['promedio'] or 0

        # Productos próximos a vencer garantía (30 días)
        from datetime import timedelta
        fecha_limite = datetime.now().date() + timedelta(days=30)
        productos_garantia_venciendo = Producto.objects.filter(
            fecha_fin_garantia__lte=fecha_limite,
            fecha_fin_garantia__gte=datetime.now().date()
        ).select_related('marca').order_by('fecha_fin_garantia')

        # Asignaciones recientes (últimos 30 días)
        fecha_limite_asignaciones = datetime.now().date() - timedelta(days=30)
        asignaciones_recientes = AsignacionHistorial.objects.filter(
            fecha_entrega__gte=fecha_limite_asignaciones,
            tipo_asignacion='ENTREGA'
        ).count()

        formato = request.GET.get('formato', 'html')

        context = {
            'fecha_reporte': datetime.now(),
            'total_productos': total_productos,
            'productos_por_estado': productos_por_estado,
            'productos_por_categoria': productos_por_categoria,
            'productos_por_departamento': productos_por_departamento,
            'valor_total_inventario': valor_total_inventario,
            'precio_promedio': precio_promedio,
            'productos_garantia_venciendo': productos_garantia_venciendo,
            'asignaciones_recientes': asignaciones_recientes,
        }

        if formato == 'pdf':
            try:
                from weasyprint import HTML, CSS
                from django.template.loader import render_to_string

                html_content = render_to_string('core/reporte_inventario.html', context)

                css = CSS(string='''
                    @page { 
                        size: A4; 
                        margin: 2cm; 
                    }
                    body { 
                        font-family: Arial, sans-serif; 
                        font-size: 12px; 
                        background: white !important;
                    }
                    h1 { 
                        color: #1a237e; 
                        page-break-after: avoid;
                    }
                    table { 
                        width: 100%; 
                        border-collapse: collapse; 
                        margin: 10px 0;
                        page-break-inside: avoid;
                    }
                    th, td { 
                        border: 1px solid #ddd; 
                        padding: 8px; 
                        text-align: left; 
                    }
                    th { 
                        background-color: #f2f2f2; 
                        font-weight: bold;
                    }
                    .section {
                        page-break-inside: avoid;
                        margin-bottom: 20px;
                    }
                    .report-header {
                        background: #1a237e !important;
                        -webkit-print-color-adjust: exact;
                        print-color-adjust: exact;
                    }
                ''')

                html = HTML(string=html_content)
                pdf = html.write_pdf(stylesheets=[css])

                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="reporte_inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

                return response

            except ImportError:
                messages.error(request, 'WeasyPrint no está instalado. Generando reporte HTML.')
                formato = 'html'

        if formato == 'html':
            return render(request, 'core/reporte_inventario.html', context)

    except Exception as e:
        messages.error(request, f'Error al generar reporte: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def eliminar_categoria(request, categoria_id):
    """Elimina definitivamente una categoría si la petición es POST"""
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, f'Categoría "{categoria.nombre}" eliminada correctamente.')
        return redirect(reverse('core:gestionar_categorias'))
    else:
        messages.error(request, 'Acción no permitida.')
        return redirect(reverse('core:gestionar_categorias'))


@login_required
def eliminar_ubicacion(request, ubicacion_id):
    ubicacion = get_object_or_404(Ubicacion, id=ubicacion_id)
    nombre = ubicacion.nombre
    ubicacion.delete()
    messages.success(request, f'Ubicación "{nombre}" eliminada correctamente.')
    return redirect('core:gestionar_ubicaciones')
